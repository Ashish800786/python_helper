from django.shortcuts import render
from urllib import request
from celery import shared_task
from email.message import EmailMessage
import smtplib
from mygeotab_project import settings
import pandas.io.sql as sql
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import colors,Alignment,Font,PatternFill,Border, Side
import uuid
from datetime import date, timedelta, datetime
import pyodbc
from decouple import config 
from django.db import connection, connections
from email_history.models import emailHistory
from pathlib import Path
from time import sleep
import os
from dateutil import tz
from pytz import timezone
import logging
from django.contrib.auth import authenticate
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse,HttpResponseRedirect
from django.contrib import messages


def excel_writer(target='',dataframe_data=None,sheet_name=None, startrow=0, startcol=0,applymap_style=None,index=False,header=False):
    with pd.ExcelWriter(target, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            print(f'Printing {sheet_name}')
            if dataframe_data is not None:
                if not dataframe_data.empty: 
                    if applymap_style:
                        dataframe_data= dataframe_data.style.applymap(applymap_style)  
                    dataframe_data.to_excel(writer, index=index, sheet_name=sheet_name, startrow=startrow, startcol=startcol, header=header)


def get_subject_name(report,func):
    subject_name=''
    if func == "1":
        if report == 'Driver Safety Scorecard 3.0 MTD':
            subject_name = f"Scheduled Report: Driver Safety Scorecard 3.0 MTD"
    
        if report == 'Asset Inspection Report':
            subject_name = f"Scheduled Report: Asset Inspection Report"
            
        if report == 'Driver Safety Scorecard 3.0':
            subject_name = f"Scheduled Report: Driver Safety Scorecard 3.0"
       
        if report == 'Fleet Distance Report':
            subject_name = f"Scheduled Report: Fleet Distance Report"

        
    if func == "2":
        if report == 'Driver Safety Scorecard 3.0':
                subject_name = f"On-Demand: Driver Safety Scorecard 3.0"
            
        if report == 'Asset Inspection Report':
            subject_name = f"On-Demand: Asset Inspection Report"
        
        if report == 'Driver Safety Scorecard 3.0 MTD':
            subject_name = f"On-Demand: Driver Safety Scorecard 3.0 MTD"

        if report == 'Fleet Distance Report':
            subject_name = f"On-Demand: Fleet Distance Report"

    return subject_name


def email_sender(emailList,msg,subject_name,scheduler_names,login_user_name,target,Start_Date,End_Date,started_Date,started_Time):
    try:
        check_valid_email=emailList.find("@")
        check_valid_dot=emailList.find(".")
        if check_valid_email == -1 or check_valid_dot == -1:
            print('========== MAIL FAILED ======')
            message= f"This email ({emailList}) is not valid email"
            email_history=emailHistory(
            fromEmail=msg["From"],
            toEmail=msg["To"],status="mail Failed",
            description=message,
            subject=subject_name,
            scheduler_name=scheduler_names,
            username=login_user_name,
            file_name=target,
            start_date=Start_Date,
            end_date=End_Date,
            started_date=started_Date,
            started_time=started_Time
            )
            email_history.save()
   
        else:
            print('========== MAIL SENDING ======')
            with smtplib.SMTP_SSL('email-smtp.us-east-1.amazonaws.com', 465) as smtp:
                smtp.login(settings.EMAIL_HOST_USER,
                        settings.EMAIL_HOST_PASSWORD)
                smtp.send_message(msg)
                message= "Report sent successfully."
                email_history=emailHistory(
                    fromEmail=msg["From"],
                    toEmail=msg["To"],status="mail sent",
                    description=message,
                    subject=subject_name,
                    scheduler_name=scheduler_names,
                    username=login_user_name,
                    file_name=target,
                    start_date=Start_Date,
                    end_date=End_Date,
                    started_date=started_Date,
                    started_time=started_Time
                    )
                email_history.save()
            
                print("Report sent successfully.")
                print('✉✉✉✉✉✉✉✉ SENDING  EMAIL✉ REPORT ✉ ✉✉✉✉✉✉✉✉')
                print('================ RECIPIENT ===============')
                print(emailList)
                print('================ SUBJECT ===============')
                print(subject_name) 
                return f"{subject_name} Report Sent.."
                
    except Exception as exception:
        print(exception)
        message= "Report sending Failed."
        email_history=emailHistory(
            fromEmail=msg["From"],
            toEmail=msg["To"],status="mail Failed",
            description=message,
            subject=subject_name,
            scheduler_name=scheduler_names,
            username=login_user_name,
            file_name=target,
            start_date=Start_Date,
            end_date=End_Date,
            started_date=started_Date,
            started_time=started_Time
            )
        email_history.save()



def required_field_checker(request,field):
    result=True
    for n in field:
        print(field.get(n))
        if field.get(n) == '' or field.get(n) == None or field.get(n) == 'None' or field.get(n) == 'null' :
            messages.error(request, f'{n.title()} field is required.')
            result=False
    return result
   
def handle_uploaded_file(f):
    with open(f'{str(settings.BASE_DIR)}/media/' + f.name,'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

            
def auth_cheker(request):
    if request.user.is_authenticated:
        print(f'USER { request.user.first_name} { request.user.last_name} IS AUTHENTICATED..')
        return True
    else:
        print('USER IS NOT AUTHENTICATED..')
        return False


def modal_objects_exception_handler(data):
    print('================ Your Modal Query Is: =========')
    print(data)
    print('============= / Your  Modal Query Is: =========')

    try:
        finall_data=data
    except Exception as e :
        finall_data=None
        print(e)

    return finall_data

def Data_History_dict_creater(current_data,updated_data,key):
        dict={
                'current_value':current_data,
                'edited_value':updated_data,
                'keyName':key,
            }
        return dict


def Data_history_creater(data):
    updated_list_String=''
    for history_data in data:
        if history_data['current_value'] != history_data['edited_value']:
                updated_list_String=updated_list_String+str((f"{history_data['keyName']}: {history_data['current_value']} , "))

    print('================ Your History Is: =========')
    print(updated_list_String)
    print('============= / Your  History Is: =========')
    return updated_list_String


def required_fields_handler(data):
    for key in data:
        if data.get(key):
          pass
        else:
            return False
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))



def data_fetch_dict_handler(query=None,django_database_name=None):
    print('================ Your Query Is: =========')
    print(query)
    print('============= / Your Query Is: =========')
    print('')
    print('================ Your Django Database Name Is: =========')
    print(django_database_name)
    print('============= / Your Django Database Name Is:=========')
    print('')
    if django_database_name == None:
        django_database_name='default'

    with connections[django_database_name].cursor() as cursor:
        try:
            cursor.execute(f"""{query}""")
            finall_data=dictfetchall(cursor)
        except Exception as e :
            finall_data={}
            print(e)
        finally:
            cursor.close()


    return finall_data



def data_fetch_DataFrame_handler(query=None,django_database_name=None):
    print('================ Your Query Is: =========')
    print(query)
    print('============= / Your Query Is: =========')
    print('')
    print('================ Your Django Database Name Is: =========')
    print(django_database_name)
    print('============= / Your Django Database Name Is:=========')
    print('')
    if django_database_name == None:
        django_database_name='default'


    with connections[django_database_name].cursor() as cursor:
        try:
            cursor.execute(f"""{query}""")
            finall_dataFrame=(pd.DataFrame(dictfetchall(cursor)))
        except Exception as e :
            finall_dataFrame=pd.DataFrame()
            print(e)
        finally:
            cursor.close()


    return finall_dataFrame


def data_cursor_dataframe(query=None,django_database_name=None):
    print('================ Your Query Is: =========')
    print(query)
    print('============= / Your Query Is: =========')
    print('')
    print('================ Your Django Database Name Is: =========')
    print(django_database_name)
    print('============= / Your Django Database Name Is:=========')
    print('')
    all_sets=[]
    if django_database_name == None:
        django_database_name='default'

    with connections[django_database_name].cursor() as cursor:
        try:
            cursor.execute(f"""{query}""")
            all_sets.append(pd.DataFrame(dictfetchall(cursor)))
            for i in range(1,11):
                second_set=cursor.nextset()
                if second_set==True:
                    print("NEXT SET")
                    all_sets.append(pd.DataFrame(dictfetchall(cursor)))

        except Exception as e :
            print(e)
        finally:
            cursor.close()
    
    print("ALL SETS ")
    print(all_sets)
    return all_sets


def data_cursor_dict_data(query=None,django_database_name=None):
    print('================ Your Query Is: =========')
    print(query)
    print('============= / Your Query Is: =========')
    print('')
    print('================ Your Django Database Name Is: =========')
    print(django_database_name)
    print('============= / Your Django Database Name Is:=========')
    print('')
    all_sets=[]
    if django_database_name == None:
        django_database_name='default'

    with connections[django_database_name].cursor() as cursor:
        try:
            cursor.execute(f"""{query}""")
            all_sets.append(dictfetchall(cursor))
            for i in range(1,11):
                second_set=cursor.nextset()
                if second_set==True:
                    print("NEXT SET")
                    all_sets.append(dictfetchall(cursor))

        except Exception as e :
            print(e)
        finally:
            cursor.close()
    
    print("ALL SETS ")
    print(all_sets)
    return all_sets

    

def data_commit_handler(query=None,django_database_name=None):
    print('================ Your Query Is: =========')
    print(query)
    print('============= / Your Query Is: =========')
    print('')
    print('================ Your Django Database Name Is: =========')
    print(django_database_name)
    print('============= / Your Django Database Name Is:=========')
    print('')
    if django_database_name == None:
        django_database_name='default'

    with connections[django_database_name].cursor() as cursor:
        try:
            cursor.execute(f"""{query}""")
            cursor.commit()
            result=True

        except Exception as e :
            result=False
            print(e)
        finally:
            cursor.close()

    return result



def data_Update_handler(data=None,tableName=None,where_condition=None,django_database_name=None):
    Update_Query=''
    comma=','
    i=0
    for key in data:
        field_data=str(data.get(key)).replace("'","''")
        if key == 'updated_at':
            Update_Query=Update_Query+f"""\n{key} = GETDATE(){comma}"""
        else:
            Update_Query=Update_Query+f"""\n{key} = '{field_data}'{comma}"""

    Update_Query=Update_Query+")"
    Update_Query=Update_Query.replace(",)","\n")
    Update_Query=Update_Query+' '+where_condition
    query=f""" UPDATE {tableName} SET \n {Update_Query} """
    result=data_commit_handler(query,django_database_name=None)
    return result



def data_Insert_handler(data=None,tableName=None,django_database_name=None):
    FieldNames='('
    values='VALUES ('
    comma=','
    i=0
    for key in data:
        field_data=str(data.get(key)).replace("'","''")
        FieldNames=FieldNames+f"""\n{key}{comma}"""
        values=values+f"""\n'{field_data}'{comma}"""
        
    values=values+")"
    FieldNames=FieldNames+")"
    FieldNames=FieldNames.replace(",)",")")
    values=values.replace("',)","'\n)")
    query=f""" INSERT INTO {tableName} {FieldNames} \n {values} """
    result=data_commit_handler(query,django_database_name=None)
    return result


def search_query_maker(data):
    i=0
    result=''
    for key in data:
        field_data=data.get(key)
        if field_data:
             result=result+f"""AND CONVERT(VARCHAR, {key}) ='{field_data}' """
    return result


def excel_downloader(target,original,filename,sheet_name,dataframe_data,start_row,start_column):
        dfAllallInventoryManagementList=dataframe_data
        print(dfAllallInventoryManagementList)
 
        BASE_DIR = os.path.dirname(
            os.path.dirname(os.path.abspath(__file__)))
        
        # coping the template file
        shutil.copyfile(original, target)

        # load excel file
        workbook = load_workbook(f'{target}')

    #---------save Workbook---------->
        workbook.save(filename=target)
    #---------/save Workbook---------->

        def all_text_align_center(val):
            return 'text-align:center;'

        print("Step 9")
        with pd.ExcelWriter(target, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            if not dfAllallInventoryManagementList.empty:
                        dfAllallInventoryManagementList= dfAllallInventoryManagementList.style.applymap(all_text_align_center) 
                        dfAllallInventoryManagementList.to_excel(
                                writer, index=False, sheet_name=f"{sheet_name}", startrow=start_row, startcol=start_column, header=False)
        return target
        # with open(target, 'rb') as fh:
        #     response = HttpResponse(
        #         fh.read(), content_type="application/vnd.ms-excel")
        #     response['Content-Disposition'] = 'inline; filename=' + \
        #         os.path.basename(target)
        #     return response
        # return True

def download_response(target):
    with open(target, 'rb') as fh:
            response = HttpResponse(
                fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + \
                os.path.basename(target)
            return response



def data_store_proc_dictData_handler(data=None,sp_name=None,django_database_name=None):
    parameters=''
    comma=','
    i=0
    for key in data:
        field_data=data.get(key)
        parameters=parameters+f"""\n{key} ='{field_data}'{comma}"""

    parameters=parameters+")"
    parameters=parameters.replace("',)","'\n")
    query=f""" EXECUTE {sp_name} \n {parameters} \n"""
    print('SP query')
    print(query)
    if django_database_name:
        database_name=django_database_name
    else:
        database_name='dupes'
    result=data_fetch_dict_handler(query,database_name)  
    return result


def data_store_proc_nextset_dataframe_handler(data=None,sp_name=None,django_database_name=None):
    parameters=''
    comma=','
    i=0
    for key in data:
        field_data=data.get(key)
        parameters=parameters+f"""\n{key} ='{field_data}'{comma}"""

    parameters=parameters+")"
    parameters=parameters.replace("',)","'\n")
    query=f""" EXECUTE {sp_name} \n {parameters} \n"""
    print('SP query')
    print(query)
    if django_database_name:
        database_name=django_database_name
    else:
        database_name='dupes'
    result=data_cursor_dataframe(query,database_name)  
    return result

def data_store_proc_nextset_dictData_handler(data=None,sp_name=None,django_database_name=None):
    parameters=''
    comma=','
    i=0
    for key in data:
        field_data=data.get(key)
        parameters=parameters+f"""\n{key} ='{field_data}'{comma}"""

    parameters=parameters+")"
    parameters=parameters.replace("',)","'\n")
    query=f""" EXECUTE {sp_name} \n {parameters} \n"""
    print('SP query')
    print(query)
    if django_database_name:
        database_name=django_database_name
    else:
        database_name='dupes'
    result=data_cursor_dict_data(query,database_name)  
    return result


def data_store_proc_query_DataFrame_handler(data=None,sp_name=None,django_database_name=None):
    parameters=''
    comma=','
    i=0
    for key in data:
        field_data=data.get(key)
        parameters=parameters+f"""\n{key} ='{field_data}'{comma}"""

    parameters=parameters+")"
    parameters=parameters.replace("',)","'\n")
    query=f""" EXECUTE {sp_name} \n {parameters} \n"""
    print('SP query')
    print(query)
    if django_database_name:
        database_name=django_database_name
    else:
        database_name='dupes'
    result=data_fetch_DataFrame_handler(query,database_name)  
    return result








#  START : COMMON FUNCTION TO SET SCHEMA OF TABLE FETCHED BY CUERSOR ===========>
def dictfetchall(cursor):
    "Return all rows from a cursor as a dict" 
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
      ]
#  END : COMMON FUNCTION TO SET SCHEMA OF TABLE FETCHED BY CUERSOR ===========>



#   START  : COMMON FUNCTION TO DATE FORMATE (M-D-Y) 01-13-2023 ===========>
def dateFormater(date,formate=None):
    if date: 
            print('================ Your Parameters Is: =========')
            print(f'================ Your Parameters Date Is: {date} =========')
            print(f'================ Your Parameters Date Formate Is: {formate} =========')
            print('============= / Your Parameters Is: =========')
            if formate == None:
                try:
                    formatedDate=datetime.strptime(str(date),"%Y-%m-%d").strftime("%m-%d-%Y")
                except Exception as e:
                    try:
                        formatedDate=datetime.strptime(str(date),"%m-%d-%Y").strftime("%m-%d-%Y")
                    except Exception as e:
                        try:
                            formatedDate=datetime.strptime(str(date),"%d-%m-%Y").strftime("%m-%d-%Y")
                        except Exception as e:
                            try:
                                formatedDate= datetime.strptime(str(date),"%m/%d/%Y").strftime("%m-%d-%Y")
                            except Exception as e:
                                try:
                                    formatedDate= datetime.strptime(str(date),"%d/%m/%Y").strftime("%m-%d-%Y")
                                except Exception as e:
                                    try:
                                        formatedDate= datetime.strptime(str(date),"%Y/%m/%d").strftime("%m-%d-%Y")
                                    except Exception as e:
                                        formatedDate=None
                                        print(e)
            else:
                try:
                    formatedDate=datetime.strptime(str(date),"%Y-%m-%d").strftime(formate)
                except Exception as e:
                    try:
                        formatedDate=datetime.strptime(str(date),"%m-%d-%Y").strftime(formate)
                    except Exception as e:
                        try:
                            formatedDate=datetime.strptime(str(date),"%d-%m-%Y").strftime(formate)
                        except Exception as e:
                            try:
                                formatedDate= datetime.strptime(str(date),"%m/%d/%Y").strftime(formate)
                            except Exception as e:
                                try:
                                    formatedDate= datetime.strptime(str(date),"%d/%m/%Y").strftime(formate)
                                except Exception as e:
                                    try:
                                        formatedDate= datetime.strptime(str(date),"%Y/%m/%d").strftime(formate)
                                    except Exception as e:
                                        formatedDate=None
                                        print(e)
            print('================ Your Final Data Is: =========')
            print(f'================ Your Final Date Is: {date} =========')
            print(f'================ Your Final Date Formate Is: {formate} =========')
            print('============= / Your Final Data Is: =========') 
    else:
         formatedDate=''          
    return formatedDate
#   END : COMMON FUNCTION TO DATE FORMATE (M-D-Y) 01-13-2023 ===========>




#   START  : COMMON FUNCTION TO DATE FORMATE (M-D-Y) 01-13-2023 ===========>
def getDateTimeYourTimeZone(TimeZoneName):
    print(f'================ Your TimeZone  Is: {date} =========')
    CurrentDateTime=datetime.now()
    print(f'================ Current UTC Date  Is: {CurrentDateTime} =========')
    try:
        TZ_DateTime=datetime.now().astimezone(timezone(TimeZoneName))
    except Exception as e:
            TZ_DateTime=None
            print(e)
    print(f'================ Your Final TimeZoneDate  Is: {date} =========')                    
    return TZ_DateTime
#   END : COMMON FUNCTION TO DATE FORMATE (M-D-Y) 01-13-2023 ===========>



#   START  : COMMON FUNCTION TO DATE FORMATE (M-D-Y) 01-13-2023 ===========>
def getDateTimeYourTimeZoneDateFormater(TimeZoneName,formate=None):
    if formate == None :
        try:
            TZ_DateTime=datetime.now().astimezone(timezone(TimeZoneName)).date().strftime("%m-%d-%Y")
        except Exception as e:
            TZ_DateTime=None
            print(e)     
    else:
        try:
            TZ_DateTime=datetime.now().astimezone(timezone(TimeZoneName)).date().strftime(formate) 
        except Exception as e:
            TZ_DateTime=None
            print(e)         
    return TZ_DateTime
#   END : COMMON FUNCTION TO DATE FORMATE (M-D-Y) 01-13-2023 ===========>


#   START  : COMMON FUNCTION TO DATE FORMATE (M-D-Y) HH:MM:SS ===========>
def getDateTimeYourTimeZoneTimeFormater(TimeZoneName,formate=None):
    if formate == None :
        try:
            TZ_DateTime=datetime.now().astimezone(timezone(TimeZoneName)).time().strftime("%H:%M:%S")
        except Exception as e:
            TZ_DateTime=None
            print(e)   
    else:
        try:
            TZ_DateTime=datetime.now().astimezone(timezone(TimeZoneName)).time().strftime(formate) 
        except Exception as e:
            TZ_DateTime=None
            print(e)        
    return TZ_DateTime
#   END : COMMON FUNCTION TO DATE FORMATE (H:M:S) HH:MM:SS ===========>


#   START : COMMON FUNCTION TO DATE FORMATE JAN 01, 2023 ===========>
def date_formate(date):
    if date :
        date=dateFormater(date,"%m-%d-%Y")
        print('date',date)
        print('====')
        date_split=str(date).split('-')
        month=''
        if date_split[0] == '01':
            month='Jan'
        if date_split[0] == '02':
            month='Feb'
        if date_split[0] == '03':
            month='Mar'
        if date_split[0] == '04':
            month='Apr'
        if date_split[0] == '05':
            month='May'
        if date_split[0] == '06':
            month='Jun'
        if date_split[0] == '07':
            month='Jul'
        if date_split[0] == '08':
            month='Aug'
        if date_split[0] == '09':
            month='Sep'
        if date_split[0] == '10':
            month='Oct'
        if date_split[0] == '11':
            month='Nov'
        if date_split[0] == '12':
            month='Dec'
        
        print('date_split: ',date_split)
        print('====')
        final_date=month+' '+date_split[1]+', '+date_split[2]
    else:
        final_date='' 
    return final_date
#  END : COMMON FUNCTION TO DATE FORMATE JAN 01, 2023 ===========>


# Excel Styling =============================================================================>
def stylingOfexcel(target_file,workbook, worksheet_name, dataframe_data, cells,color=None,stylemapColor=None):
    workbook = load_workbook(filename=target_file)
    wb = workbook[worksheet_name]
    if not dataframe_data.empty:
        #-----Left Align----------->
        for i in range(9, len(dataframe_data)+9):  
            Align_left = Alignment(horizontal='left')
            if "left" in cells:
                for j in cells["left"]:
                    wb[j+str(i)].alignment=Align_left
        #-----/Left Align--------->
                    
                    
        #-----Right Align----------->
        for i in range(9, len(dataframe_data)+9) :  
            Align_right = Alignment(horizontal='right')
            if "right" in cells:
                for j in cells["right"]:
                    wb[j+str(i)].alignment=Align_right
        #-----/Right Align--------->
        

        #-----Right Align----------->
        for i in range(9, len(dataframe_data)+9) :  
            Align_center = Alignment(horizontal='center')
            if "center" in cells:
                for j in cells["center"]:
                    wb[j+str(i)].alignment=Align_center
        #-----/Center Align--------->                
        
        
        #-----border Align----------->
        for i in range(9, len(dataframe_data)+9) :  
            #----- / Cell Border Colors ------------------------------------------->
            CellBorder = Side(border_style="thick", color="000000")
            #----- / Cell Border Colors ------------------------------------------->
            
            
            if "border" in cells:
                for j in cells["border"]:
                    wb[j+str(i)].border=Border(top=CellBorder, left=CellBorder, right=CellBorder, bottom=CellBorder)
        #-----/Center Align--------->                
        
        #-----BackGround Colors -------------------------------------------- >
            bgBlue = PatternFill(start_color='25477B', end_color='25477B', fill_type = "solid")
            FontWhite = Font(color = "FFFFFF",bold = True)
            # bgBlue = PatternFill(bgColor="25477B", fill_type = "solid")
            bgGreen = PatternFill(start_color='7AFF33', end_color='7AFF33', fill_type = "solid")
            bgYellow = PatternFill(start_color='FFEF97', end_color='FFEF97', fill_type = "solid")
            bgBrown = PatternFill(start_color='FDD7A4', end_color='FDD7A4',fill_type = "solid")
            bgRed = PatternFill(start_color='FF3333', end_color='FF3333', fill_type = "solid")
        #----- / BackGround Colors ------------------------------------------->
        
        
        # Background Color -------------------------------------------------------------------->   
        if color =='red' :
             for i in range(9, len(dataframe_data)+9) :  
                if "color" in cells:
                    for j in cells["color"]:
                        wb[j+str(i)].fill=bgRed 
        
        if color =='green' :
             for i in range(9, len(dataframe_data)+9) :  
                if "color" in cells:
                    for j in cells["color"]:
                        wb[j+str(i)].fill=bgGreen 
                   
            #----- / Cell Border Colors ------------------------------------------->
            
            #----- / Cell Border Colors ------------------------------------------->
            # dataframe_data= dataframe_data.style.applymap(stylemapColor, subset=[4])

        # / Background Color -------------------------------------------------------------------->   


    # ---------save Workbook---------->
    workbook.save(filename=target_file)
    # --------- / save Workbook---------->
                         
    return 
# / Excel Styling ==================================================================================>






